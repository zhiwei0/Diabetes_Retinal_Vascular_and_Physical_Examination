# Test_9 实验专用，重新做眼底血管数据的预测，采用单因素+多因素回归分析
import pandas as pd
import os
import statsmodels.api as sm
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font
from Analysis_and_Prediction import Prediction_Models
from Prediction2.Test_8 import add_value_to_excel_column1


# 获得列名
def get_cols_name():
    input_file_path = 'D:/Analysis of diabetes/Test_9/data1.xlsx'
    # 此处获取待归一化的指标来赋值归一化中的cols_to_normalize值
    # 读取Excel文件，指定文件路径和sheet_name参数（如果sheet在多个工作表中有相同的名称，可以使用sheet_name参数指定具体的sheet）
    df = pd.read_excel(input_file_path)

    # 获取列名列表
    column_names = df.columns.tolist()

    print(column_names)


# 处理数据
def process_data(input_file_path, output_file_path, cols_1):
    # 读取整个Excel文件中的所有表格
    xls = pd.ExcelFile(input_file_path)
    sheets = xls.sheet_names

    i = 0

    with pd.ExcelWriter(output_file_path) as writer:
        for sheet_name in sheets:
            df = xls.parse(sheet_name)

            # 删除指定列
            df.drop(columns=cols_1, inplace=True, errors='ignore')

            # 删除包含0的数据行
            df = df[~(df == 0).any(axis=1)]

            # 删除存在空值的数据行
            df.dropna(inplace=True)

            # 删除非数值行
            cols_to_check = ['artery_caliber', 'vein_caliber', 'frac', 'AVR', 'artery_curvature',
                         'vein_curvature', 'artery_BSTD', 'vein_BSTD', 'artery_simple_curvatures',
                         'vein_simple_curvatures', 'artery_Branching_Coefficient', 'vein_Branching_Coefficient',
                         'artery_Num1stBa', 'vein_Num1stBa', 'artery_Branching_Angle', 'vein_Branching_Angle',
                         'artery_Angle_Asymmetry', 'vein_Angle_Asymmetry', 'artery_Asymmetry_Raito',
                         'vein_Asymmetry_Raito', 'artery_JED', 'vein_JED', 'artery_Optimal_Deviation',
                         'vein_Optimal_Deviation', 'vessel_length_density', 'vessel_area_density']
            for col in cols_to_check:
                df[col] = pd.to_numeric(df[col], errors='coerce')  # 将非数值数据转换为NaN
            df = df.dropna(subset=cols_to_check)  # 删除含有NaN的行

            # 增加新列Result
            col_name = df.columns.tolist()
            index = col_name.index('ID') + 1
            col_name.insert(index, 'Result')
            df = df.reindex(columns=col_name)
            # 对新列赋值
            df['Result'] = i
            i += 1

            # 保存
            df.to_excel(writer, sheet_name=sheet_name, index=False)


    # 两两合并表格
    df_dict = pd.read_excel(output_file_path, sheet_name=None)

    # 合并sheet1和sheet2为sheet01
    sheet01 = pd.concat([df_dict['正常范围'], df_dict['中间范围']])

    # 合并sheet2和sheet3为sheet12
    sheet12 = pd.concat([df_dict['中间范围'], df_dict['糖尿病']])

    # 保存到新的Excel文件
    with pd.ExcelWriter(output_file_path) as writer:
        sheet01.to_excel(writer, sheet_name='正常范围_中间范围', index=False)
        sheet12.to_excel(writer, sheet_name='中间范围_糖尿病', index=False)


    sheet_name = '中间范围_糖尿病'  # 指定要处理的表格名称
    # 全局的替换规则字典，可以使用嵌套字典或者字典列表
    # 使用嵌套字典的例子：
    # 进行数值替换时无需加''
    column_replace_dict = {
        'Result': {1: 0, 2: 1}
    }

    # 读取指定表格数据
    df = pd.read_excel(output_file_path, sheet_name=sheet_name)

    # 针对每列进行值替换（使用全局的替换规则字典）
    for column, value_map in column_replace_dict.items():
        if column in df.columns:
            df[column] = df[column].replace(value_map)

    # 保存修改后的数据到原文件中
    with pd.ExcelWriter(output_file_path, mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("处理完成并保存为新的Excel文件。")


def data_normalization(input_file_path,output_file_path):
    df_0_1 = pd.read_excel(input_file_path, sheet_name='正常范围_中间范围')
    df_1_2 = pd.read_excel(input_file_path, sheet_name='中间范围_糖尿病1')

    # 获取原值减去均值除以方差的特征列
    cols_to_normalize = ['artery_caliber', 'vein_caliber', 'frac', 'AVR', 'artery_curvature',
                         'vein_curvature', 'artery_BSTD', 'vein_BSTD', 'artery_simple_curvatures',
                         'vein_simple_curvatures', 'artery_Branching_Coefficient', 'vein_Branching_Coefficient',
                         'artery_Num1stBa', 'vein_Num1stBa', 'artery_Branching_Angle', 'vein_Branching_Angle',
                         'artery_Angle_Asymmetry', 'vein_Angle_Asymmetry', 'artery_Asymmetry_Raito',
                         'vein_Asymmetry_Raito', 'artery_JED', 'vein_JED', 'artery_Optimal_Deviation',
                         'vein_Optimal_Deviation', 'vessel_length_density', 'vessel_area_density']
    df_0_1[cols_to_normalize] = df_0_1[cols_to_normalize].apply(lambda x: (x - x.mean()) / x.std())
    df_1_2[cols_to_normalize] = df_1_2[cols_to_normalize].apply(lambda x: (x - x.mean()) / x.std())

    # 保存修改后的数据列至新的 Excel 文件
    with pd.ExcelWriter(output_file_path) as writer:
        df_0_1.to_excel(writer, sheet_name='正常范围_中间范围', index=False)
        df_1_2.to_excel(writer, sheet_name='中间范围_糖尿病', index=False)

    print("特征归一化处理并保存为新的Excel文件。")


# 单因素回归分析
def univariate_regression(input_file_path, output_file_path):
    # 读取Excel文件
    file_path = input_file_path

    # 获取Excel文件中的所有表格名
    xls = pd.ExcelFile(file_path)
    sheet_names = xls.sheet_names

    # 创建新的Excel文件来保存结果
    output_file = output_file_path

    output_workbook = pd.ExcelWriter(output_file, engine='openpyxl')

    # 遍历每张表格进行处理
    for sheet_name in sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # 拆分数据为自变量和因变量
        medical_data = df.loc[:, df.columns[2:]]
        y = df.iloc[:, 1]

        # Initialize an empty DataFrame to store the results for the current sheet
        results_df = pd.DataFrame()

        # 遍历每一列进行单因素回归分析
        for col in medical_data.columns:
            # 获取目标变量与预测变量
            X = medical_data[col]

            # 添加常数项截距
            X = sm.add_constant(X)

            # 构建逻辑回归模型
            model = sm.Logit(y, X)
            result = model.fit()
            # result =  model.fit_regularized(alpha=0,method='l1')
            params = result.params
            conf_int = result.conf_int()
            or_vals = np.exp(params)
            or_vals.name = "OR"
            conf_int = np.exp(conf_int)
            p_values = result.pvalues
            p_values.name = "P值"

            # 合并当前列的结果到一个DataFrame中，并按每个自变量一行的形式输出
            summary = pd.concat([or_vals, conf_int, p_values], axis=1)
            summary.columns = ["OR", "95% CI Lower", "95% CI Upper", "P值"]
            summary.index = X.columns

            # Exclude the 'const' row from the summary
            summary_without_const = summary.iloc[1:]

            # 添加当前列的结果到当前表格结果DataFrame中
            results_df = pd.concat([results_df, summary_without_const], axis=0)

        # 将当前表格的结果保存到新的Excel文件的当前表中，并设置样式
        results_df.to_excel(output_workbook, sheet_name=sheet_name, index=True)

    # 保存结果到新的Excel文件
    output_workbook.save()

    print("单因素回归分析完成")


# 多因素回归分析(两张表格）
def multivariate_regression(input_file_path, output_file_path, col_1, col_2):
    # 读取Excel文件
    file_path = input_file_path
    output_file = output_file_path

    # 打开Excel文件
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames

    # 定义第一张表格和第二张表格的自变量列名列表
    X_columns_sheet1 = col_1
    X_columns_sheet2 = col_2

    # 遍历每张表格进行多因素回归分析并保存结果
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name in sheet_names:
            # 读取数据
            df = pd.read_excel(file_path, sheet_name=sheet_name)

            # 拆分数据为自变量和因变量
            y_column = "Result"  # 设置因变量的列名
            X_columns = X_columns_sheet1 if sheet_name == "正常范围_中间范围" else X_columns_sheet2  # 根据表格选择自变量列名列表
            X = df[X_columns]
            y = df[y_column]

            # 添加常数项截距
            X = sm.add_constant(X)

            # 多因素逻辑回归模型拟合
            model = sm.Logit(y, X)
            results = model.fit()
            # results = model.fit_regularized(alpha=0, method='l1')
            params = results.params
            conf_int = results.conf_int()
            or_vals = np.exp(params)
            or_vals.name = "OR"
            conf_int = np.exp(conf_int)
            p_values = results.pvalues
            p_values.name = "P值"

            # 合并所有结果到一个DataFrame中，并按每个自变量一行的形式输出
            summary = pd.concat([or_vals, conf_int, p_values], axis=1)
            summary.columns = ["OR", "95% CI Lower", "95% CI Upper", "P值"]
            summary.index = X.columns

            # Exclude the 'const' row from the summary
            summary_without_const = summary.iloc[1:]

            results_df = pd.DataFrame()
            results_df = pd.concat([results_df, summary_without_const], axis=0)

            results_df.to_excel(writer, sheet_name=sheet_name, index=True)

    print("多因素回归分析完成")


# 多因素回归分析 单张表格分析
def multivariate_regression_1(input_file_path, output_file_path, col_1, sheet_name):
    # 读取Excel文件
    file_path = input_file_path
    output_file = output_file_path

    # 定义第一张表格自变量列名列表
    X_columns_sheet1 = col_1

    # 遍历每张表格进行多因素回归分析并保存结果
    with pd.ExcelWriter(output_file) as writer:
        # 读取数据
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # 拆分数据为自变量和因变量
        y_column = "Result"  # 设置因变量的列名
        X_columns = X_columns_sheet1
        X = df[X_columns]
        y = df[y_column]

        # 添加常数项截距
        X = sm.add_constant(X)

        # 多因素逻辑回归模型拟合
        model = sm.Logit(y, X)
        results = model.fit()
        # results = model.fit_regularized(alpha=0, method='l1')
        params = results.params
        conf_int = results.conf_int()
        or_vals = np.exp(params)
        or_vals.name = "OR"
        conf_int = np.exp(conf_int)
        p_values = results.pvalues
        p_values.name = "P值"

        # 合并所有结果到一个DataFrame中，并按每个自变量一行的形式输出
        summary = pd.concat([or_vals, conf_int, p_values], axis=1)
        summary.columns = ["OR", "95% CI Lower", "95% CI Upper", "P值"]
        summary.index = X.columns

        # Exclude the 'const' row from the summary
        summary_without_const = summary.iloc[1:]

        results_df = pd.DataFrame()
        results_df = pd.concat([results_df, summary_without_const], axis=0)

        results_df.to_excel(writer, sheet_name=sheet_name, index=True)

    print("单张表格的多因素回归分析完成")


# 标记指标
def high_light(input_file_path, output_file_path):
    file_path = input_file_path
    output_file = output_file_path

    # 打开Excel文件
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames

    # 定义标红的阈值
    threshold = 0.05

    # 用于存储所有E列数值小于0.05的行对应的第一列的值
    first_column_values = []

    # 遍历每张表格进行处理
    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        # 获取表格中指定列的列号（假设需要处理的列为E列）
        column_to_check = 'E'
        column_idx = ws[column_to_check]

        # 遍历每行数据
        for row in column_idx:
            # 检查单元格的值是否小于阈值
            if isinstance(row.value, (int, float)) and row.value < threshold:
                # 将满足条件的单元格设置为加粗
                cell = ws.cell(row=row.row, column=row.column)
                bold_font = Font(color="FF0000", bold=True)
                cell.font = bold_font

                # 获取对应的A列单元格，并设置字体颜色为红色
                cell_a = ws.cell(row=row.row, column=1)  # 假设第一列为A列
                red_font = Font(color="FF0000", bold=True)
                cell_a.font = red_font

                # 将该行对应的第一列的值添加到列表中
                first_column_values.append(cell_a.value)

    # 保存修改后的Excel文件
    wb.save(output_file)

    print("P值小于0.05字体加粗并A列字体颜色为红色完成")
    print("小于0.05的指标:", first_column_values)


def main():
    # 预处理
    input_file_path = 'D:/Analysis of diabetes/Test_9/data0.xlsx'
    output_file_path = 'D:/Analysis of diabetes/Test_9/data1.xlsx'
    cols_1 = ['artery_LDR', 'vein_LDR']
    process_data(input_file_path, output_file_path, cols_1)

    # 特征归一化
    input_file_path = 'D:/Analysis of diabetes/Test_9/data1.xlsx'
    output_file_path = 'D:/Analysis of diabetes/Test_9/data_normalization.xlsx'
    data_normalization(input_file_path, output_file_path)

    # 单因素回归分析
    input_file_path = 'D:/Analysis of diabetes/Test_9/data_normalization.xlsx'
    output_file_path = 'D:/Analysis of diabetes/Test_9/result_univariate.xlsx'
    univariate_regression(input_file_path, output_file_path)


    # 筛选P<0.05
    input_file_path = 'D:/Analysis of diabetes/Test_9/result_univariate.xlsx'
    output_file_path = 'D:/Analysis of diabetes/Test_9/result_univariate_h.xlsx'
    high_light(input_file_path, output_file_path)


# 多因素回归分析
def multi():
    input_file_path = 'D:/Analysis of diabetes/Test_9/data_normalization.xlsx'
    output_file_path = 'D:/Analysis of diabetes/Test_9/result_multivariate.xlsx'
    col_1 = []
    col_2 = ['frac', 'vein_curvature', 'vein_Branching_Angle', 'vessel_length_density', 'vessel_area_density']
    multivariate_regression(input_file_path, output_file_path, col_1, col_2)

    input_file_path = 'D:/Analysis of diabetes/Test_9/result_multivariate.xlsx'
    output_file_path = 'D:/Analysis of diabetes/Test_9/result_multivariate_h.xlsx'
    high_light(input_file_path, output_file_path)


# 多因素回归分析
def multi1():
    input_file_path = 'D:/Analysis of diabetes/Test_9/data_normalization.xlsx'
    output_file_path = 'D:/Analysis of diabetes/Test_9/result_multivariate_1.xlsx'
    sheet_name = '中间范围_糖尿病' #设置为需要分析的表格名
    col_1 = ['frac', 'vein_curvature', 'vein_Branching_Angle', 'vessel_length_density', 'vessel_area_density']
    multivariate_regression_1(input_file_path, output_file_path, col_1, sheet_name)

    input_file_path = 'D:/Analysis of diabetes/Test_9/result_multivariate_1.xlsx'
    output_file_path = 'D:/Analysis of diabetes/Test_9/result_multivariate_1_h.xlsx'
    high_light(input_file_path, output_file_path)


# 代入模型预测
def prediction():
    input_file_path = "D:/Analysis of diabetes/Test_9/data_normalization.xlsx"
    sheet_name = '中间范围_糖尿病'
    output_file_path = "D:/Analysis of diabetes/Test_9/预测模型结果.xlsx"
    pic_path = "D:/Analysis of diabetes/Test_9/ROC/"
    cols_x = ['frac', 'vein_Branching_Angle', 'vessel_length_density']
    cols_y = 'Result'

    columns_values = {'cols_x': cols_x}
    add_value_to_excel_column1(output_file_path, 'Sheet1', columns_values)

    Prediction_Models.Prediction_Models(input_file_path, sheet_name, output_file_path, pic_path, cols_x, cols_y)


def prediction1():
    # 使用单因素回归分析筛出的指标
    input_file_path = "D:/Analysis of diabetes/Test_9/data_normalization.xlsx"
    sheet_name = '中间范围_糖尿病'
    output_file_path = "D:/Analysis of diabetes/Test_9/预测模型结果.xlsx"
    pic_path = "D:/Analysis of diabetes/Test_9/ROC/"
    cols_x = ['frac', 'vein_curvature', 'vein_Branching_Angle', 'vessel_length_density', 'vessel_area_density']
    cols_y = 'Result'

    columns_values = {'cols_x': cols_x}
    add_value_to_excel_column1(output_file_path, 'Sheet1', columns_values)

    Prediction_Models.Prediction_Models(input_file_path, sheet_name, output_file_path, pic_path, cols_x, cols_y)


def prediction2():
    input_file_path = "D:/Analysis of diabetes/Test_9/data_normalization.xlsx"
    sheet_names = ['正常范围_中间范围', '中间范围_糖尿病']
    output_file_path = "D:/Analysis of diabetes/Test_9/预测模型结果.xlsx"
    pic_path = "D:/Analysis of diabetes/Test_9/ROC/"
    cols_x = ['artery_caliber', 'vein_caliber', 'frac', 'AVR', 'artery_curvature',
             'vein_curvature', 'artery_BSTD', 'vein_BSTD', 'artery_simple_curvatures',
             'vein_simple_curvatures', 'artery_Branching_Coefficient', 'vein_Branching_Coefficient',
             'artery_Num1stBa', 'vein_Num1stBa', 'artery_Branching_Angle', 'vein_Branching_Angle',
             'artery_Angle_Asymmetry', 'vein_Angle_Asymmetry', 'artery_Asymmetry_Raito',
             'vein_Asymmetry_Raito', 'artery_JED', 'vein_JED', 'artery_Optimal_Deviation',
             'vein_Optimal_Deviation', 'vessel_length_density', 'vessel_area_density']
    cols_y = 'Result'

    for sheet_name in sheet_names:
        columns_values = {'sheet_name': sheet_name,
                          'cols_x': 'ALL',
                          }
        add_value_to_excel_column1(output_file_path, 'Sheet1', columns_values)

        Prediction_Models.Prediction_Models(input_file_path, sheet_name, output_file_path, pic_path, cols_x, cols_y)


if __name__ == '__main__':
    # main()
    # multi1()
    # prediction()
    prediction1()
    # prediction2()