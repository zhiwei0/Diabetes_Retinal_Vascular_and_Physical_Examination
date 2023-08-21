import pandas as pd
import numpy as np
import statsmodels.api as sm
from Analysis_and_Prediction import Prediction_Models
from Prediction2.Test_8 import add_value_to_excel_column1
from openpyxl import load_workbook
from openpyxl.styles import Font

# 获取列名
def get_cols_name(input_file_path):
    df = pd.read_excel(input_file_path)
    columns_name = df.columns.tolist()
    print(columns_name)

# 处理数据
def process_data(input_file_path, output_file_path, cols_1):
    # 读取整个Excel文件中的所有表格
    xls = pd.ExcelFile(input_file_path)
    sheets = xls.sheet_names

    i = 0

    with pd.ExcelWriter(output_file_path) as writer:
        for sheet_name in sheets:
            df = xls.parse(sheet_name)

            # 删除指定列(XM, result、血糖、糖化血红蛋白）
            df.drop(columns=cols_1, inplace=True, errors='ignore')

            # 删除包含0的数据行
            columns_to_check = ['年龄', '体重', '身高', '舒张压', '收缩压', '总蛋白(TP)', '白蛋白(Alb)', '球蛋白(GLB)', '总胆红素(T-BIL)', '直接胆红素(DB)', '间接胆红素(IB)', '谷丙转氨酶(ALT)', '谷草转氨酶(AST)', '尿素氮(BUN)', '肌酐(Cr)', '尿酸(UA)', '总胆固醇（TC）', '甘油三酯（TG）', '高密度脂蛋白胆固醇', '低密度脂蛋白胆固醇']
            df = df[~(df[columns_to_check] == 0).any(axis=1)]
            # df = df[~(df == 0).any(axis=1)]

            # 删除存在空值的数据行
            df.dropna(inplace=True)

            # 删除非数值行
            # 设置要检查的列
            cols_to_check = ['年龄', '性别', '体重', '身高', '舒张压', '收缩压', '总蛋白(TP)', '白蛋白(Alb)', '球蛋白(GLB)', '总胆红素(T-BIL)', '直接胆红素(DB)', '间接胆红素(IB)', '谷丙转氨酶(ALT)', '谷草转氨酶(AST)', '尿素氮(BUN)', '肌酐(Cr)', '尿酸(UA)', '总胆固醇（TC）', '甘油三酯（TG）', '高密度脂蛋白胆固醇', '低密度脂蛋白胆固醇']
            for col in cols_to_check:
                df[col] = pd.to_numeric(df[col], errors='coerce')  # 将非数值数据转换为NaN
            df = df.dropna(subset=cols_to_check)  # 删除含有NaN的行

            # 增加新列Result
            col_name = df.columns.tolist()
            index = col_name.index('ID') + 1
            col_name.insert(index, 'Result')
            df = df.reindex(columns=col_name)
            # 对新列赋值
            df['Result'] = i
            i += 1

            # 保存
            df.to_excel(writer, sheet_name=sheet_name, index=False)


    # 两两合并表格
    df_dict = pd.read_excel(output_file_path, sheet_name=None)

    # 合并sheet1和sheet2为sheet01
    sheet01 = pd.concat([df_dict['正常范围'], df_dict['中间范围']])

    # 合并sheet2和sheet3为sheet12
    sheet12 = pd.concat([df_dict['中间范围'], df_dict['糖尿病']])

    # 保存到新的Excel文件
    with pd.ExcelWriter(output_file_path) as writer:
        sheet01.to_excel(writer, sheet_name='正常范围_中间范围', index=False)
        sheet12.to_excel(writer, sheet_name='中间范围_糖尿病', index=False)


    sheet_name = '中间范围_糖尿病'  # 指定要处理的表格名称
    # 全局的替换规则字典，可以使用嵌套字典或者字典列表
    # 使用嵌套字典的例子：
    # 进行数值替换时无需加''
    column_replace_dict = {
        'Result': {1: 0, 2: 1}
    }

    # 读取指定表格数据
    df = pd.read_excel(output_file_path, sheet_name=sheet_name)

    # 针对每列进行值替换（使用全局的替换规则字典）
    for column, value_map in column_replace_dict.items():
        if column in df.columns:
            df[column] = df[column].replace(value_map)

    # 保存修改后的数据到原文件中
    with pd.ExcelWriter(output_file_path, mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("处理完成并保存为新的Excel文件。")


# Glu和PRO进行值映射 再计算BMI
# 定义计算BMI的函数
def calculate_bmi(df):
    df['身高（m）'] = df['身高'] / 100  # 将厘米转换为米
    df['BMI'] = df['体重'] / (df['身高（m）'] * df['身高（m）'])

    # 删除指定若干列，可以将需要删除的列名放在这个列表中
    columns_to_delete = ['身高（m）', '身高', '体重']
    df.drop(columns=columns_to_delete, inplace=True, errors='ignore')
    return df

def process_Data1(input_file_path, output_file_path):
    # 读取Excel文件，指定文件路径和sheet_name参数（如果sheet在多个工作表中有相同的名称，可以使用sheet_name参数指定具体的sheet）
    file_path = input_file_path
    output_file = output_file_path

    df_dict = pd.read_excel(file_path, sheet_name=None)

    # 针对每张表格，指定要替换的列和替换的映射关系
    columns_to_replace = {'尿糖（Glu）': {'阴性': 0, '+-': 1, '1+': 1, '2+': 1, '3+': 1, '4+': 1},
                          '蛋白PRO': {'阴性': 0, '+-': 1, '1+': 1, '2+': 1, '3+': 1, '4+': 1}}

    # 处理每张表格
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, df in df_dict.items():
            # 针对每列进行值替换
            for column, replace_dict in columns_to_replace.items():
                df[column] = df[column].replace(replace_dict)

            # 计算BMI
            df = calculate_bmi(df)

            # 保存到新的Excel文件
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("值映射完成，BMI计算完成并已保存到新的Excel文件。")


# 特征归一化
def data_normalization(input_file_path,output_file_path):
    df_0_1 = pd.read_excel(input_file_path, sheet_name='正常范围_中间范围')
    df_1_2 = pd.read_excel(input_file_path, sheet_name='中间范围_糖尿病')

    # 获取原值减去均值除以方差的特征列
    cols_to_normalize = ['年龄', '性别', '舒张压', '收缩压', '尿糖（Glu）', '蛋白PRO', '总蛋白(TP)', '白蛋白(Alb)', '球蛋白(GLB)', '总胆红素(T-BIL)', '直接胆红素(DB)', '间接胆红素(IB)', '谷丙转氨酶(ALT)', '谷草转氨酶(AST)', '尿素氮(BUN)', '肌酐(Cr)', '尿酸(UA)', '总胆固醇（TC）', '甘油三酯（TG）', '高密度脂蛋白胆固醇', '低密度脂蛋白胆固醇', 'BMI']
    df_0_1[cols_to_normalize] = df_0_1[cols_to_normalize].apply(lambda x: (x - x.mean()) / x.std())
    df_1_2[cols_to_normalize] = df_1_2[cols_to_normalize].apply(lambda x: (x - x.mean()) / x.std())

    # 保存修改后的数据列至新的 Excel 文件
    with pd.ExcelWriter(output_file_path) as writer:
        df_0_1.to_excel(writer, sheet_name='正常范围_中间范围', index=False)
        df_1_2.to_excel(writer, sheet_name='中间范围_糖尿病', index=False)

    print("特征归一化处理并保存到"+output_file_path)


# 将眼底血管数据和体检数据合并
def merge_data_by_ID(input_file_path1, input_file_path2, output_file_path):
    # 读取文件1中的表1和文件2中的表2
    xls = pd.ExcelFile(input_file_path1)
    sheets = xls.sheet_names

    with pd.ExcelWriter(output_file_path) as writer:
        for sheet_name in sheets:
            df1 = pd.read_excel(input_file_path1, sheet_name=sheet_name)
            df2 = pd.read_excel(input_file_path2, sheet_name=sheet_name)

            # 根据ID和NO列进行匹配，并拼接成新的数据行
            merged_data = pd.merge(df1, df2, left_on='NO', right_on='ID', suffixes=('_file1', '_file2'))

            # 将结果保存到新的表格中
            merged_data.to_excel(writer, sheet_name=sheet_name, index=False)

    print("合并完成并保存到" + output_file_path)


# 分别对每张表格的数据进行单因素回归分析
def univariate_regression_analysis(input_file_path, output_file_path):
    # 读取Excel文件
    file_path = input_file_path

    # 获取Excel文件中的所有表格名
    xls = pd.ExcelFile(file_path)
    sheet_names = xls.sheet_names

    # 创建新的Excel文件来保存结果
    output_file = output_file_path

    output_workbook = pd.ExcelWriter(output_file, engine='openpyxl')

    # 遍历每张表格进行处理
    for sheet_name in sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # 拆分数据为自变量和因变量
        medical_data = df.loc[:, df.columns[2:]]
        y = df.iloc[:, 1]

        # Initialize an empty DataFrame to store the results for the current sheet
        results_df = pd.DataFrame()

        # 遍历每一列进行单因素回归分析
        for col in medical_data.columns:
            # 获取目标变量与预测变量
            X = medical_data[col]

            # 添加常数项截距
            X = sm.add_constant(X)

            # 构建逻辑回归模型
            model = sm.Logit(y, X)
            result = model.fit()
            # result =  model.fit_regularized(alpha=0,method='l1')
            params = result.params
            conf_int = result.conf_int()
            or_vals = np.exp(params)
            or_vals.name = "OR"
            conf_int = np.exp(conf_int)
            p_values = result.pvalues
            p_values.name = "P值"

            # 合并当前列的结果到一个DataFrame中，并按每个自变量一行的形式输出
            summary = pd.concat([or_vals, conf_int, p_values], axis=1)
            summary.columns = ["OR", "95% CI Lower", "95% CI Upper", "P值"]
            summary.index = X.columns

            # Exclude the 'const' row from the summary
            summary_without_const = summary.iloc[1:]

            # 添加当前列的结果到当前表格结果DataFrame中
            results_df = pd.concat([results_df, summary_without_const], axis=0)

        # 将当前表格的结果保存到新的Excel文件的当前表中，并设置样式
        results_df.to_excel(output_workbook, sheet_name=sheet_name, index=True)

    # 保存结果到新的Excel文件
    output_workbook.save()

    print("已经完成单因素回归分析,并将结果保存到: "+output_file_path)


# 分别对每张表格进行多因素回归分析
def multivariate_regression_analysis(input_file_path, output_file_path):
    xls = pd.ExcelFile(input_file_path)
    sheet_names = xls.sheet_names

    # 每张表格在单因素回归分析筛选出的指标列名列表
    X_columns_sheet1 = ['vein_caliber_L1', 'artery_simple_curvatures_L1', 'artery_Optimal_Deviation_L1', 'artery_Asymmetry_Raito_L2', '年龄', '收缩压', '白蛋白(Alb)', '球蛋白(GLB)']
    X_columns_sheet2 = ['frac_R1', 'vessel_length_density_R1', 'vessel_area_density_R1', 'vein_Optimal_Deviation_R2', 'vessel_area_density_R2', '舒张压', '蛋白PRO', '甘油三酯（TG）']

    with pd.ExcelWriter(output_file_path) as writer:
        for sheet_name in sheet_names:
            # 读取数据
            df = pd.read_excel(input_file_path, sheet_name= sheet_name)

            # 拆分数据为自变量和因变量
            y_column = "Result" # 设置因变量的列名
            # 根据表格名选择自变量的列名列表
            X_columns = X_columns_sheet1 if sheet_name == "正常范围_中间范围" else X_columns_sheet2
            X = df[X_columns]
            y = df[y_column]

            # 创建一个空表格存储数据
            results_df = pd.DataFrame()

            # 添加常数项截距
            X = sm.add_constant(X)

            # 多因素逻辑回归模型拟合
            model = sm.Logit(y, X)
            results = model.fit()
            params = results.params
            or_vals = np.exp(params)
            or_vals.name = "OR"
            conf_int = results.conf_int()
            conf_int = np.exp(conf_int)
            p_values = results.pvalues
            p_values.name = "P值"

            # 合并所有结果到一个DataFrame中,并按每个自变量一行的形式输出
            summary = pd.concat([or_vals,conf_int,p_values], axis=1)
            summary.columns = ["OR", "95% CI Lower", "95% CI Upper", "P值"]
            summary.index = X.columns

            # 将const排除
            summary_without_const = summary.iloc[1:]

            results_df = pd.concat([results_df, summary_without_const], axis=0)
            results_df.to_excel(writer, sheet_name= sheet_name, index=True)
        # 将结果保存到新的Excel文件中
        writer.save()
    print("已完成多因素回归分析,并将结果保存到: "+output_file_path)


# 把结果中P值小于0.05的加粗/标红
def highlight_result(input_file_path, output_file_path):
    file_path = input_file_path
    output_file = output_file_path

    # 打开Excel文件
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames

    # 定义标红的阈值
    threshold = 0.05

    # 遍历每张表格进行处理
    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        # 获取表格中指定列的列号（假设需要处理的列为E列）
        column_to_check = 'E'

        column_idx = ws[column_to_check]

        # 遍历每行数据
        for row in column_idx:
            # 检查单元格的值是否小于阈值
            if isinstance(row.value, (int, float)) and row.value < threshold:
                # 将满足条件的单元格设置为加粗
                cell = ws.cell(row=row.row, column=row.column)
                bold_font = Font(color="FF0000", bold=True)
                cell.font = bold_font

                # 获取对应的A列单元格，并设置字体颜色为红色
                cell_a = ws.cell(row=row.row, column=1)  # 假设第一列为A列
                red_font = Font(color="FF0000", bold=True)
                cell_a.font = red_font

    # 保存修改后的Excel文件
    wb.save(output_file)

    print("P值小于0.05字体加粗并A列字体颜色为红色完成")


def highlight_result_pro(input_file_path, output_file_path):
    file_path = input_file_path
    output_file = output_file_path

    # 打开Excel文件
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames

    # 定义标红的阈值
    threshold = 0.05

    # 遍历每张表格进行处理
    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        # 获取表格中指定列的列号（假设需要处理的列为E列）
        column_to_check = 'E'
        column_idx = ws[column_to_check]

        # 数组用于存储符合条件的第一列的值
        collected_values = []

        # 遍历每行数据
        for row in column_idx:
            # 检查单元格的值是否小于阈值
            if isinstance(row.value, (int, float)) and row.value < threshold:
                # 将满足条件的单元格设置为加粗
                cell = ws.cell(row=row.row, column=row.column)
                bold_font = Font(color="FF0000", bold=True)
                cell.font = bold_font

                # 获取对应的A列单元格，并设置字体颜色为红色
                cell_a = ws.cell(row=row.row, column=1)  # 假设第一列为A列
                red_font = Font(color="FF0000", bold=True)
                cell_a.font = red_font

                # 收集符合条件的第一列的值
                collected_values.append(cell_a.value)

        # 将收集的值存储到该工作表的最后一行
        ws.append(collected_values)
        print(f"For sheet {sheet_name}, collected values are: {collected_values}")

    # 保存修改后的Excel文件
    wb.save(output_file)

    print("P值小于0.05字体加粗并A列字体颜色为红色完成")


# 所有指标代入
def prediction():
    input_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_retinal_phy.xlsx'
    output_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/预测模型结果.xlsx'
    sheet_names = ['正常范围_中间范围', '中间范围_糖尿病']
    pic_path = "D:/Analysis of diabetes/Test_5/ROC/"
    cols_x = ['artery_caliber_R1', 'vein_caliber_R1', 'frac_R1', 'AVR_R1', 'artery_curvature_R1', 'vein_curvature_R1', 'artery_BSTD_R1', 'vein_BSTD_R1', 'artery_simple_curvatures_R1', 'vein_simple_curvatures_R1', 'artery_Branching_Coefficient_R1', 'vein_Branching_Coefficient_R1', 'artery_Num1stBa_R1', 'vein_Num1stBa_R1', 'artery_Branching_Angle_R1', 'vein_Branching_Angle_R1', 'artery_Angle_Asymmetry_R1', 'vein_Angle_Asymmetry_R1', 'artery_Asymmetry_Raito_R1', 'vein_Asymmetry_Raito_R1', 'artery_JED_R1', 'vein_JED_R1', 'artery_Optimal_Deviation_R1', 'vein_Optimal_Deviation_R1', 'vessel_length_density_R1', 'vessel_area_density_R1',
          'artery_caliber_L1', 'vein_caliber_L1', 'frac_L1', 'AVR_L1', 'artery_curvature_L1', 'vein_curvature_L1', 'artery_BSTD_L1', 'vein_BSTD_L1', 'artery_simple_curvatures_L1', 'vein_simple_curvatures_L1', 'artery_Branching_Coefficient_L1', 'vein_Branching_Coefficient_L1', 'artery_Num1stBa_L1', 'vein_Num1stBa_L1', 'artery_Branching_Angle_L1', 'vein_Branching_Angle_L1', 'artery_Angle_Asymmetry_L1', 'vein_Angle_Asymmetry_L1', 'artery_Asymmetry_Raito_L1', 'vein_Asymmetry_Raito_L1', 'artery_JED_L1', 'vein_JED_L1', 'artery_Optimal_Deviation_L1', 'vein_Optimal_Deviation_L1', 'vessel_length_density_L1', 'vessel_area_density_L1',
          'artery_caliber_R2', 'vein_caliber_R2', 'frac_R2', 'AVR_R2', 'artery_curvature_R2', 'vein_curvature_R2', 'artery_BSTD_R2', 'vein_BSTD_R2', 'artery_simple_curvatures_R2', 'vein_simple_curvatures_R2', 'artery_Branching_Coefficient_R2', 'vein_Branching_Coefficient_R2', 'artery_Num1stBa_R2', 'vein_Num1stBa_R2', 'artery_Branching_Angle_R2', 'vein_Branching_Angle_R2', 'artery_Angle_Asymmetry_R2', 'vein_Angle_Asymmetry_R2', 'artery_Asymmetry_Raito_R2', 'vein_Asymmetry_Raito_R2', 'artery_JED_R2', 'vein_JED_R2', 'artery_Optimal_Deviation_R2', 'vein_Optimal_Deviation_R2', 'vessel_length_density_R2', 'vessel_area_density_R2',
          'artery_caliber_L2', 'vein_caliber_L2', 'frac_L2', 'AVR_L2', 'artery_curvature_L2', 'vein_curvature_L2', 'artery_BSTD_L2', 'vein_BSTD_L2', 'artery_simple_curvatures_L2', 'vein_simple_curvatures_L2', 'artery_Branching_Coefficient_L2', 'vein_Branching_Coefficient_L2', 'artery_Num1stBa_L2', 'vein_Num1stBa_L2', 'artery_Branching_Angle_L2', 'vein_Branching_Angle_L2', 'artery_Angle_Asymmetry_L2', 'vein_Angle_Asymmetry_L2', 'artery_Asymmetry_Raito_L2', 'vein_Asymmetry_Raito_L2', 'artery_JED_L2', 'vein_JED_L2', 'artery_Optimal_Deviation_L2', 'vein_Optimal_Deviation_L2', 'vessel_length_density_L2', 'vessel_area_density_L2',
          '年龄', '性别', 'BMI', '舒张压', '收缩压', '尿糖（Glu）', '蛋白PRO', '总蛋白(TP)', '白蛋白(Alb)', '球蛋白(GLB)', '总胆红素(T-BIL)', '直接胆红素(DB)', '间接胆红素(IB)', '谷丙转氨酶(ALT)', '谷草转氨酶(AST)', '尿素氮(BUN)', '肌酐(Cr)', '尿酸(UA)', '总胆固醇（TC）', '甘油三酯（TG）', '高密度脂蛋白胆固醇', '低密度脂蛋白胆固醇']

    cols_y = 'Result'

    add_value_to_excel_column1(output_file_path, 'Sheet1', '所有指标')

    for sheet_name in sheet_names:
        columns_values = {'cols_x': cols_x,
                          'sheet_name': sheet_name
                          }
        add_value_to_excel_column1(output_file_path, 'Sheet1', columns_values)

        Prediction_Models.Prediction_Models(input_file_path, sheet_name, output_file_path, pic_path, cols_x, cols_y)

    print("所有指标代入预测完成")


# 使用单因素回归分析筛出的指标
def prediction1():
    input_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_retinal_phy.xlsx'
    output_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/预测模型结果.xlsx'
    sheet_names = ['正常范围_中间范围', '中间范围_糖尿病']
    pic_path = "D:/Analysis of diabetes/Test_5/ROC/"
    cols_x_01 = ['vein_caliber_L1', 'artery_simple_curvatures_L1', 'artery_Optimal_Deviation_L1', 'artery_Asymmetry_Raito_L2', '年龄', '收缩压', '白蛋白(Alb)', '球蛋白(GLB)']
    cols_x_12 = ['frac_R1', 'vessel_length_density_R1', 'vessel_area_density_R1', 'vein_Optimal_Deviation_R2', 'vessel_area_density_R2', '舒张压', '蛋白PRO', '甘油三酯（TG）']
    cols = [cols_x_01, cols_x_12]
    cols_y = 'Result'

    add_value_to_excel_column1(output_file_path, 'Sheet1', '单因素')

    for sheet_name, cols_x in zip(sheet_names, cols):

        columns_values = {
                          'sheet_name': sheet_name,
                          'cols_x': cols_x
                          }
        add_value_to_excel_column1(output_file_path, 'Sheet1', columns_values)

        Prediction_Models.Prediction_Models(input_file_path, sheet_name, output_file_path, pic_path, cols_x, cols_y)

    print("单因素回归分析指标代入预测完成")


# 使用多因素回归分析筛出的指标
def prediction2():
    input_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_retinal_phy.xlsx'
    output_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/预测模型结果.xlsx'
    sheet_names = ['正常范围_中间范围', '中间范围_糖尿病']
    pic_path = "D:/Analysis of diabetes/Test_5/ROC/"
    cols_x_01 = ['vein_caliber_L1', 'artery_simple_curvatures_L1', 'artery_Optimal_Deviation_L1', '收缩压']
    cols_x_12 = ['vein_Optimal_Deviation_R2', '甘油三酯（TG）']
    cols = [cols_x_01, cols_x_12]
    cols_y = 'Result'

    add_value_to_excel_column1(output_file_path, 'Sheet1', '多因素')

    for sheet_name, cols_x in zip(sheet_names, cols):

        columns_values = {
                          'sheet_name': sheet_name,
                          'cols_x': cols_x
                          }
        add_value_to_excel_column1(output_file_path, 'Sheet1', columns_values)

        Prediction_Models.Prediction_Models(input_file_path, sheet_name, output_file_path, pic_path, cols_x, cols_y)

    print("多因素回归分析指标代入预测完成")


def prePreocss():
    # 1获得列名
    input_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_p2.xlsx'
    get_cols_name(input_file_path)

    # 2预处理
    # input_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_p.xlsx'
    # output_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_p1.xlsx'
    # # 要删除的列
    # cols_1 = ['XM', 'result', '尿糖', '血糖', '糖化血红蛋白']
    # process_data(input_file_path, output_file_path, cols_1)

    # 3值映射+BMI计算
    # input_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_p1.xlsx'
    # output_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_p2.xlsx'
    # process_Data1(input_file_path, output_file_path)

    # # 4特征归一化
    # input_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_p2.xlsx'
    # output_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_normal_phy.xlsx'
    # data_normalization(input_file_path, output_file_path)

    # # 5 将体检数据和长眼底血管数据拼接
    # input_file_path1 = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_normal_retinal.xlsx'
    # input_file_path2 = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_normal_phy.xlsx'
    # output_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_retinal_phy.xlsx'
    # merge_data_by_ID(input_file_path1, input_file_path2, output_file_path)


def analysis_data():
    # 1 单因素回归分析
    input_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/data_retinal_phy.xlsx'
    # output_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/result_univariate.xlsx'
    # univariate_regression_analysis(input_file_path, output_file_path)


    output_file_path = 'D:/Analysis of diabetes/Test_5/第六次实验数据/result_multivariate.xlsx'
    multivariate_regression_analysis(input_file_path, output_file_path)

    highlight_result_pro(output_file_path, output_file_path)


def main():
    # prediction()
    prediction1()
    prediction2()


if __name__ == '__main__':
    # prePreocss()
    # analysis_data()
    main()