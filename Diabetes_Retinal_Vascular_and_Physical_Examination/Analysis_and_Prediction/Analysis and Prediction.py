import pandas as pd
import numpy as np
import statsmodels.api as sm
from openpyxl.styles import Font
from openpyxl import load_workbook

def data_normalization(input_file_path, output_file_path):
    # 输入的文件应该是两种状态的数据合并后的，然后进行特征归一化
    # 归一化方法为：原值减去均值除以方差

    # 读取表格中的数据
    df_0_1 = pd.read_excel(input_file_path, sheet_name="正常范围_中间范围")
    df_1_2 = pd.read_excel(input_file_path, sheet_name="中间范围_糖尿病")

    # 要进行特征归一化的列
    cols_to_normalize = ['artery_caliber_R1', 'vein_caliber_R1', 'frac_R1', 'AVR_R1', 'artery_curvature_R1', 'vein_curvature_R1', 'artery_BSTD_R1', 'vein_BSTD_R1', 'artery_simple_curvatures_R1', 'vein_simple_curvatures_R1', 'artery_Branching_Coefficient_R1', 'vein_Branching_Coefficient_R1', 'artery_Num1stBa_R1', 'vein_Num1stBa_R1', 'artery_Branching_Angle_R1', 'vein_Branching_Angle_R1', 'artery_Angle_Asymmetry_R1', 'vein_Angle_Asymmetry_R1', 'artery_Asymmetry_Raito_R1', 'vein_Asymmetry_Raito_R1', 'artery_JED_R1', 'vein_JED_R1', 'artery_Optimal_Deviation_R1', 'vein_Optimal_Deviation_R1', 'vessel_length_density_R1', 'vessel_area_density_R1',
                         'artery_caliber_L1', 'vein_caliber_L1', 'frac_L1', 'AVR_L1', 'artery_curvature_L1', 'vein_curvature_L1', 'artery_BSTD_L1', 'vein_BSTD_L1', 'artery_simple_curvatures_L1', 'vein_simple_curvatures_L1', 'artery_Branching_Coefficient_L1', 'vein_Branching_Coefficient_L1', 'artery_Num1stBa_L1', 'vein_Num1stBa_L1', 'artery_Branching_Angle_L1', 'vein_Branching_Angle_L1', 'artery_Angle_Asymmetry_L1', 'vein_Angle_Asymmetry_L1', 'artery_Asymmetry_Raito_L1', 'vein_Asymmetry_Raito_L1', 'artery_JED_L1', 'vein_JED_L1', 'artery_Optimal_Deviation_L1', 'vein_Optimal_Deviation_L1', 'vessel_length_density_L1', 'vessel_area_density_L1',
                         'artery_caliber_R2', 'vein_caliber_R2', 'frac_R2', 'AVR_R2', 'artery_curvature_R2', 'vein_curvature_R2', 'artery_BSTD_R2', 'vein_BSTD_R2', 'artery_simple_curvatures_R2', 'vein_simple_curvatures_R2', 'artery_Branching_Coefficient_R2', 'vein_Branching_Coefficient_R2', 'artery_Num1stBa_R2', 'vein_Num1stBa_R2', 'artery_Branching_Angle_R2', 'vein_Branching_Angle_R2', 'artery_Angle_Asymmetry_R2', 'vein_Angle_Asymmetry_R2', 'artery_Asymmetry_Raito_R2', 'vein_Asymmetry_Raito_R2', 'artery_JED_R2', 'vein_JED_R2', 'artery_Optimal_Deviation_R2', 'vein_Optimal_Deviation_R2', 'vessel_length_density_R2', 'vessel_area_density_R2',
                         'artery_caliber_L2', 'vein_caliber_L2', 'frac_L2', 'AVR_L2', 'artery_curvature_L2', 'vein_curvature_L2', 'artery_BSTD_L2', 'vein_BSTD_L2', 'artery_simple_curvatures_L2', 'vein_simple_curvatures_L2', 'artery_Branching_Coefficient_L2', 'vein_Branching_Coefficient_L2', 'artery_Num1stBa_L2', 'vein_Num1stBa_L2', 'artery_Branching_Angle_L2', 'vein_Branching_Angle_L2', 'artery_Angle_Asymmetry_L2', 'vein_Angle_Asymmetry_L2', 'artery_Asymmetry_Raito_L2', 'vein_Asymmetry_Raito_L2', 'artery_JED_L2', 'vein_JED_L2', 'artery_Optimal_Deviation_L2', 'vein_Optimal_Deviation_L2', 'vessel_length_density_L2', 'vessel_area_density_L2']

    # 进行特征归一化
    df_0_1[cols_to_normalize] = df_0_1[cols_to_normalize].apply(lambda x:(x - x.mean()) / x.std())
    df_1_2[cols_to_normalize] = df_1_2[cols_to_normalize].apply(lambda x:(x - x.mean()) / x.std())

    with pd.ExcelWriter(output_file_path) as writer:
        df_0_1.to_excel(writer, sheet_name="正常范围_中间范围", index = False)
        df_1_2.to_excel(writer, sheet_name="中间范围_糖尿病", index = False)

    print("原数据进行特征归一化完成,已保存到: "+output_file_path)

def univariate_regression_analysis(input_file_path, output_file_path):
    # 分别对每张表格的数据进行单因素回归分析

    xls = pd.ExcelFile(input_file_path)
    sheet_names = xls.sheet_names

    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        # 遍历每张表格进行处理
        for sheet_name in sheet_names:
            # 读取数据
            df = pd.read_excel(input_file_path, sheet_name= sheet_name)

            # 修改位需要进行分析的列
            medical_data = df.loc[:, df.columns[2:106]]

            y = df.iloc[:, 1]

            # 创建一个空表格存储数据
            results_df = pd.DataFrame()

            # 遍历每一列进行单因素回归分析
            for col in medical_data.columns:
                X = medical_data[col]

                # 添加常数项截距
                X = sm.add_constant(X)

                # 构建逻辑回归模型
                model = sm.Logit(y, X)
                result = model.fit()
                params = result.params
                conf_int = result.conf_int()
                or_vals = np.exp(params)
                or_vals.name = "OR"
                conf_int = np.exp(conf_int)
                p_values = result.pvalues
                p_values.name = "P值"

                # 合并当前列的结果到一个DataFrame中,并按每个自变量一行的形式输出
                summary = pd.concat([or_vals, conf_int, p_values], axis=1)
                summary.columns = ["OR", "95% CI Lower", "95% CI Upper", "P值"]
                summary.index = X.columns

                # 将const行去除
                summary_without_const = summary.iloc[1:]

                # 将当前列结果增加到当前表格结果DataFrame中
                results_df = pd.concat([results_df, summary_without_const], axis=0)

            # 将当前表格的结果保存到新的Excel文件的当前表格中,并设置样式
            results_df.to_excel(writer, sheet_name= sheet_name, index=True)
        # 结果保存到新的Excel文件中
        writer.save()

    print("已经完成单因素回归分析,并将结果保存到: "+output_file_path)


def multivariate_regression_analysis(input_file_path, output_file_path):
    # 分别对每张表格进行多因素回归分析

    xls = pd.ExcelFile(input_file_path)
    sheet_names = xls.sheet_names

    # 每张表格在单因素回归分析筛选出的指标列名列表
    X_columns_sheet1 = []
    X_columns_sheet2 = []

    with pd.ExcelWriter(output_file_path) as writer:
        for sheet_name in sheet_names:
            # 读取数据
            df = pd.read_excel(input_file_path, sheet_name= sheet_name)

            # 拆分数据为自变量和因变量
            y_column = "Result" # 设置因变量的列名
            # 根据表格名选择自变量的列名列表
            X_columns = X_columns_sheet1 if sheet_name == "正常范围_中间范围" else X_columns_sheet2
            X = df[X_columns]
            y = df[y_column]

            # 创建一个空表格存储数据
            results_df = pd.DataFrame()

            # 添加常数项截距
            X = sm.add_constant(X)

            # 多因素逻辑回归模型拟合
            model = sm.Logit(y, X)
            results = model.fit()
            params = results.params
            or_vals = np.exp(params)
            or_vals.name = "OR"
            conf_int = results.conf_int()
            conf_int = np.exp(conf_int)
            p_values = results.pvalues
            p_values.name = "P值"

            # 合并所有结果到一个DataFrame中,并按每个自变量一行的形式输出
            summary = pd.concat([or_vals,conf_int,p_values], axis=1)
            summary.columns = ["OR", "95% CI Lower", "95% CI Upper", "P值"]
            summary.index = X.columns

            # 将const排除
            summary_without_const = summary.iloc[1:]

            results_df = pd.concat([results_df, summary_without_const], axis=0)
            results_df.to_excel(writer, sheet_name= sheet_name, index=True)
        # 将结果保存到新的Excel文件中
        writer.save()
    print("已完成多因素回归分析,并将结果保存到: "+output_file_path)


def highlight_result(input_file_path, output_file_path):
    # 把结果中P值小于0.05的加粗/标红
    threshold = 0.05

    # xls = pd.ExcelFile(input_file_path)
    xls = load_workbook(input_file_path)
    sheet_names = xls.sheet_names

    for sheet_name in sheet_names:
        df = pd.read_excel(input_file_path, sheet_name=sheet_name)

        # 获取表格中指定列的列号(假设需要处理的列为E列)
        column_to_check= "E"
        column_idx = df[column_to_check]

        # 遍历检查每行数据
        for row in column_idx:
            # 检查单元格的值是否小于阈值
            if isinstance(row.value, (int, float)) and row.value < threshold:
                # 将满足条件的单元格设置为加粗
                cell = df.cell(row=row.row, column= row.column)
                bold_font = Font(color="FF0000", bold= True)
                cell.font = bold_font

                # 设置为字体红色
                # 假设第一列为A列
                cell_a = df.cell(row=row.row, column=1)
                red_font = Font(color="FF0000", bold=True)
                cell_a.font = red_font
    # 保存修改后的Excel文件
    xls.save(output_file_path)
    print("P值小于0.05字体加粗并A列字体颜色为红色完成，并保存到"+output_file_path)


if __name__ == '__main__':
    # input_file_path = "D:/Analysis of diabetes/Test_5/data_combine.xlsx"
    output_file_path1 = "D:/Analysis of diabetes/Test_5/data_normalization.xlsx"
    # data_normalization(input_file_path, output_file_path1)

    output_file_path2 = "D:/Analysis of diabetes/Test_5/result_univariate.xlsx"
    univariate_regression_analysis(output_file_path1,output_file_path2)

