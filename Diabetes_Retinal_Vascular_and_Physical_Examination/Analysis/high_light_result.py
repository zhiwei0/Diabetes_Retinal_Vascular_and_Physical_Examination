from openpyxl import load_workbook
from openpyxl.styles import Font

# 读取Excel文件
file_path = "D:/Analysis of diabetes/Test_5/result_merge.xlsx"
output_file = "D:/Analysis of diabetes/Test_5/result_merge_high_light.xlsx"

# 打开Excel文件
wb = load_workbook(file_path)
sheet_names = wb.sheetnames

# 定义标红的阈值
threshold = 0.05

# 遍历每张表格进行处理
for sheet_name in sheet_names:
    ws = wb[sheet_name]

    # 获取表格中指定列的列号（假设需要处理的列为E，I列）
    column_indices = [5, 9]  # Column E and I (Index starts from 1)

    # 遍历每行数据
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        for col_idx in column_indices:
            cell_value = row[col_idx - 1]
            if cell_value is not None and isinstance(cell_value, (int, float)) and cell_value < threshold:
                # 将满足条件的单元格设置为加粗
                cell = ws.cell(row=row_idx, column=col_idx)
                bold_font = Font(color="FF0000", bold=True)
                cell.font = bold_font

                # 获取对应的A列单元格，并设置字体颜色为红色
                cell_a = ws.cell(row=row_idx, column=1)  # 假设第一列为A列
                red_font = Font(color="FF0000", bold=True)
                cell_a.font = red_font

# 保存修改后的Excel文件
wb.save(output_file)

print("P值小于0.05字体加粗并A列字体颜色为红色完成")
