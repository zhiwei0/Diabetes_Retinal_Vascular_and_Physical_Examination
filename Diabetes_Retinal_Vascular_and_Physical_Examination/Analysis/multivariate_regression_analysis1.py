# 可循环执行对每张表格的多因素回归分析
import pandas as pd
import numpy as np
import statsmodels.api as sm
from openpyxl import load_workbook

# 读取Excel文件
file_path = "D:/Analysis of diabetes/Test_5/data_normalization.xlsx"
output_file = "D:/Analysis of diabetes/Test_5/result_multivariate.xlsx"

# 打开Excel文件
wb = load_workbook(file_path)
sheet_names = wb.sheetnames

# 定义第一张表格和第二张表格的自变量列名列表
X_columns_sheet1 = ["vein_Branching_Angle_R1", "artery_simple_curvatures_L1", "artery_Optimal_Deviation_L1",
                    "vein_Asymmetry_Raito_R2", "artery_Asymmetry_Raito_L2"]
X_columns_sheet2 = ["vessel_length_density_R1", "vessel_area_density_R1", "vein_Optimal_Deviation_R2",
                    "frac_L2", "vessel_area_density_L2"]

# 遍历每张表格进行多因素回归分析并保存结果
with pd.ExcelWriter(output_file) as writer:
    for sheet_name in sheet_names:
        # 读取数据
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # 拆分数据为自变量和因变量
        y_column = "Result"  # 设置因变量的列名
        X_columns = X_columns_sheet1 if sheet_name == "正常范围_中间范围" else X_columns_sheet2  # 根据表格选择自变量列名列表
        X = df[X_columns]
        y = df[y_column]

        # 添加常数项截距
        X = sm.add_constant(X)

        # 多因素逻辑回归模型拟合
        model = sm.Logit(y, X)
        results = model.fit()
        # results = model.fit_regularized(alpha=0, method='l1')
        params = results.params
        conf_int = results.conf_int()
        or_vals = np.exp(params)
        or_vals.name = "OR"
        conf_int = np.exp(conf_int)
        p_values = results.pvalues
        p_values.name = "P值"

        # 合并所有结果到一个DataFrame中，并按每个自变量一行的形式输出
        summary = pd.concat([or_vals, conf_int, p_values], axis=1)
        summary.columns = ["OR", "95% CI Lower", "95% CI Upper", "P值"]
        summary.index = X.columns

        # Exclude the 'const' row from the summary
        summary_without_const = summary.iloc[1:]

        results_df = pd.DataFrame()
        results_df = pd.concat([results_df, summary_without_const],axis=0)

        results_df.to_excel(writer,sheet_name=sheet_name,index=True)

print("Regression results exported to Excel successfully.")
